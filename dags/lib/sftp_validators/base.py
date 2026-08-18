"""
Base framework cho SFTP validators.

Thiết kế:
  - `ValidationResult`: dataclass capture errors/warnings + metrics
  - `SFTPValidator`: abstract base, subclass override validate_batch()
  - Helper methods share giữa các validator (connect SFTP, peek file,
    count CSV columns, inspect xlsx sheets, match content keywords)

Mở rộng:
  - Thêm pipeline mới → tạo file mới subclass SFTPValidator
  - Thêm check type mới → thêm helper method vào base (giữ backward compat)
  - Config rules → attributes trên subclass (REQUIRED_FILES dict, MIN_COLS, …)

Dependencies:
  - paramiko (đã có trong Airflow worker)
  - openpyxl (đã có — aml_v2 parsers dùng)
"""

from __future__ import annotations

import io
import logging
import re
import zipfile
from dataclasses import dataclass, field
from typing import Optional

import paramiko
from airflow.models import Variable

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# Result container
# ══════════════════════════════════════════════════════════════════════════════
@dataclass
class ValidationResult:
    """
    Kết quả của 1 lần validate.

    - errors: list các lỗi HARD — sẽ block pipeline
    - warnings: list các cảnh báo SOFT — log nhưng vẫn cho qua
    - metrics: dict free-form để record số liệu (files_checked, rows_found, ...)
    """

    passed: bool = True
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    metrics: dict = field(default_factory=dict)

    def fail(self, msg: str) -> None:
        self.passed = False
        self.errors.append(msg)
        logger.error("[validator] FAIL: %s", msg)

    def warn(self, msg: str) -> None:
        self.warnings.append(msg)
        logger.warning("[validator] WARN: %s", msg)

    def set_metric(self, key: str, value) -> None:
        self.metrics[key] = value

    def summary(self) -> str:
        parts = [f"passed={self.passed}"]
        if self.metrics:
            parts.append(f"metrics={self.metrics}")
        if self.errors:
            parts.append(f"errors={self.errors}")
        if self.warnings:
            parts.append(f"warnings={self.warnings}")
        return " | ".join(parts)


# ══════════════════════════════════════════════════════════════════════════════
# Base validator
# ══════════════════════════════════════════════════════════════════════════════
class SFTPValidator:
    """
    Base class cho mọi pipeline validator.

    Subclass phải:
      1. Override `REMOTE_DIR_VAR` — tên Airflow Variable chứa SFTP root path
         (ví dụ 'SFTP_REMOTE_DIR_CREDIT')
      2. Override `validate_batch(sftp, result)` — logic check cụ thể

    Subclass có thể override:
      - `NAME` — tên hiển thị trong log (default: class name)
      - `__init__` để thêm params (gọi super().__init__() đầu tiên)
    """

    # Subclass override ──────────────────────────────────────────────────────
    NAME: str = "SFTPValidator"
    REMOTE_DIR_VAR: str = "SFTP_REMOTE_DIR"

    def __init__(self, cob: Optional[str] = None, month: Optional[str] = None):
        """
        cob   : YYYY-MM-DD — dùng cho pipeline daily (Pattern A)
        month : YYYY-MM — dùng cho pipeline monthly (HNF)
        Pattern B (accounting, bnk) có thể dùng cob làm target biz_date nếu muốn
        validate file match COB cụ thể, hoặc None để chỉ check general.
        """
        self.cob = cob
        self.month = month
        self.log = logging.getLogger(f"validator.{self.NAME}")

    # ── Entry point ─────────────────────────────────────────────────────────
    def run(self) -> ValidationResult:
        result = ValidationResult()
        self.log.info("=" * 60)
        self.log.info(
            f"[{self.NAME}] Start validation (cob={self.cob}, month={self.month})"
        )
        try:
            with self._connect() as sftp_ctx:
                self.validate_batch(sftp_ctx, result)
        except paramiko.SSHException as e:
            result.fail(f"SFTP connection error: {e}")
        except Exception as e:
            result.fail(f"Unexpected error: {type(e).__name__}: {e}")
        self.log.info(f"[{self.NAME}] {result.summary()}")
        self.log.info("=" * 60)
        return result

    # ── Subclass hook ───────────────────────────────────────────────────────
    def validate_batch(self, sftp: "SFTPContext", result: ValidationResult) -> None:
        raise NotImplementedError("Subclass must implement validate_batch()")

    # ── SFTP connection ─────────────────────────────────────────────────────
    def _connect(self) -> "SFTPContext":
        """
        Mở SFTP connection dùng credentials từ Airflow Variables — consistent
        với sync Spark scripts (SFTP_HOST/PORT/USER/PASS).
        """
        return SFTPContext(
            host=Variable.get("SFTP_HOST"),
            port=int(Variable.get("SFTP_PORT", default_var="22")),
            user=Variable.get("SFTP_USER"),
            password=Variable.get("SFTP_PASS"),
            remote_dir=Variable.get(self.REMOTE_DIR_VAR),
        )


# ══════════════════════════════════════════════════════════════════════════════
# SFTP context manager + helpers
# ══════════════════════════════════════════════════════════════════════════════
class SFTPContext:
    """
    Wrapper quanh paramiko SFTPClient cung cấp helper methods thường dùng.
    Dùng làm context manager để auto-close.
    """

    def __init__(self, host, port, user, password, remote_dir):
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.remote_dir = remote_dir.rstrip("/")
        self._ssh: Optional[paramiko.SSHClient] = None
        self._sftp: Optional[paramiko.SFTPClient] = None

    def __enter__(self):
        self._ssh = paramiko.SSHClient()
        self._ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self._ssh.connect(
            self.host,
            self.port,
            self.user,
            self.password,
            timeout=300,
            auth_timeout=300,
            banner_timeout=300,
            look_for_keys=False,
            allow_agent=False,
        )
        self._sftp = self._ssh.open_sftp()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            if self._sftp:
                self._sftp.close()
            if self._ssh:
                self._ssh.close()
        except Exception:
            pass

    # ── Folder / file listing ──────────────────────────────────────────────
    def list_folder(self, relative_path: str = "") -> list[paramiko.SFTPAttributes]:
        """List file attrs trong folder. relative_path rỗng = root remote_dir."""
        path = self._resolve(relative_path)
        try:
            return self._sftp.listdir_attr(path)
        except FileNotFoundError:
            return []

    def folder_exists(self, relative_path: str) -> bool:
        path = self._resolve(relative_path)
        try:
            attr = self._sftp.stat(path)
            import stat as stat_mod
            return stat_mod.S_ISDIR(attr.st_mode)
        except FileNotFoundError:
            return False

    def file_exists(self, relative_path: str) -> bool:
        path = self._resolve(relative_path)
        try:
            self._sftp.stat(path)
            return True
        except FileNotFoundError:
            return False

    def stat_file(self, relative_path: str) -> Optional[paramiko.SFTPAttributes]:
        path = self._resolve(relative_path)
        try:
            return self._sftp.stat(path)
        except FileNotFoundError:
            return None

    def read_bytes(self, relative_path: str, size: int = 20000) -> Optional[bytes]:
        """Đọc `size` byte đầu file. Dùng để inspect header/sample."""
        path = self._resolve(relative_path)
        try:
            with self._sftp.open(path, "rb") as fh:
                return fh.read(size)
        except FileNotFoundError:
            return None

    def read_full(self, relative_path: str) -> Optional[bytes]:
        """Đọc toàn bộ file — DÙNG CẨN THẬN cho file nhỏ (xlsx < 20MB)."""
        path = self._resolve(relative_path)
        try:
            with self._sftp.open(path, "rb") as fh:
                return fh.read()
        except FileNotFoundError:
            return None

    def _resolve(self, relative_path: str) -> str:
        if not relative_path:
            return self.remote_dir
        return f"{self.remote_dir}/{relative_path.lstrip('/')}"


# ══════════════════════════════════════════════════════════════════════════════
# Shared inspection helpers — có thể dùng trong mọi validator
# ══════════════════════════════════════════════════════════════════════════════

def detect_encoding(sample: bytes) -> str:
    """UTF-16 BOM → 'utf-16', else 'utf-8-sig' (handle UTF-8 BOM)."""
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"
    return "utf-8-sig"


def decode_sample(sample: bytes) -> str:
    """Decode sample bytes → str, auto-detect UTF-8/UTF-16."""
    if not sample:
        return ""
    enc = detect_encoding(sample)
    return sample.decode(enc, errors="ignore")


def count_csv_columns_no_header(
    sample: bytes,
    delimiter: str = "|",
) -> int:
    """
    Đếm số cột ở dòng đầu tiên của file CSV không có header.
    Dùng để verify schema match với số lượng header trong parser script.
    """
    text = decode_sample(sample)
    first_line = text.split("\n", 1)[0].rstrip("\r")
    if not first_line:
        return 0
    return first_line.count(delimiter) + 1


def get_csv_header(
    sample: bytes,
    delimiter: str = "~",
) -> list[str]:
    """
    Lấy header row từ file CSV có header. Trả về list tên cột (stripped, upper).
    """
    text = decode_sample(sample)
    first_line = text.split("\n", 1)[0].rstrip("\r")
    if not first_line:
        return []
    return [c.strip().upper() for c in first_line.split(delimiter)]


def content_contains(sample: bytes, pattern: str, ignore_case: bool = True) -> bool:
    """Check chuỗi `pattern` có trong sample không."""
    text = decode_sample(sample)
    if ignore_case:
        return pattern.upper() in text.upper()
    return pattern in text


def count_keywords(sample: bytes, keywords: list[str], ignore_case: bool = True) -> int:
    """Đếm số keyword trong sample (dùng cho validate file format T24)."""
    text = decode_sample(sample)
    if ignore_case:
        text = text.upper()
        keywords = [k.upper() for k in keywords]
    return sum(1 for kw in keywords if kw in text)


def extract_biz_date_accounting(sample: bytes) -> Optional[str]:
    """
    Extract biz_date từ header CRF T24: "AS AT CLOSE OF dd MMM yyyy".
    Trả về 'YYYY-MM-DD' hoặc None.
    """
    from datetime import datetime
    text = decode_sample(sample).upper()
    m = re.search(r"AS AT CLOSE OF\s+(\d{1,2})\s+([A-Z]{3})\s+(\d{4})", text)
    if not m:
        return None
    day, mon, year = m.groups()
    try:
        return datetime.strptime(f"{day} {mon} {year}", "%d %b %Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def extract_biz_date_bnk(sample: bytes) -> Optional[str]:
    """
    Extract MIS_DATE từ row 1 của BNK CSV (delimiter ~).
    Format: YYYYMMDD trong column value → 'YYYY-MM-DD' hoặc None.
    """
    import csv
    from datetime import datetime

    text = decode_sample(sample)
    try:
        reader = csv.DictReader(io.StringIO(text), delimiter="~")
        if not reader.fieldnames:
            return None
        reader.fieldnames = [f.strip().upper() for f in reader.fieldnames]
        first = next(reader, None)
        if not first:
            return None
        for col, val in first.items():
            if "MIS_DATE" in col:
                clean = re.sub(r"[^0-9]", "", str(val).strip())
                if len(clean) == 8:
                    try:
                        return datetime.strptime(clean, "%Y%m%d").strftime("%Y-%m-%d")
                    except ValueError:
                        return None
    except Exception:
        return None
    return None


# ── XLSX inspection (lightweight, dùng zipfile thay openpyxl cho speed) ──────

def xlsx_sheet_names(xlsx_bytes: bytes) -> list[str]:
    """
    Đọc tên các sheet trong file xlsx bằng cách parse workbook.xml.
    Nhanh hơn openpyxl nhiều lần khi chỉ cần metadata.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
            with z.open("xl/workbook.xml") as f:
                wb_xml = f.read().decode("utf-8", errors="ignore")
        return re.findall(r'<sheet[^/]*name="([^"]+)"[^/]*/>', wb_xml)
    except Exception:
        return []


def xlsx_sheet_column_count(xlsx_bytes: bytes, sheet_name: str) -> Optional[int]:
    """
    Đếm số cột (max column index) của 1 sheet cụ thể trong xlsx.
    Dùng openpyxl để inspect chính xác (handle merged cells, formulas, …).
    Return None nếu sheet không tồn tại hoặc lỗi.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), data_only=True, read_only=True
        )
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        return ws.max_column
    except Exception as e:
        logger.warning(f"openpyxl inspect failed for sheet {sheet_name}: {e}")
        return None


def xlsx_first_sheet_column_count(xlsx_bytes: bytes) -> Optional[int]:
    """Đếm số cột của sheet đầu tiên (active)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), data_only=True, read_only=True
        )
        ws = wb.active
        return ws.max_column
    except Exception as e:
        logger.warning(f"openpyxl inspect failed: {e}")
        return None
